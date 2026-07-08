# -*- coding: utf-8 -*-
"""
sync_bcra.py — baja las series monetarias/cambiarias del BCRA (API v4) y las REDUCE a los
JSON que consume el Monitor "Bull's Eye". Corre en GitHub Actions (tiene salida a internet;
el sandbox de Cowork NO llega al BCRA). Deja los archivos en macro/*.json para que el reporte
los baje como bcra_series/monetario/monetario_ext/reservas/reservas_ext/usd_cer/bm_componentes.

Ademas baja los .xlsx publicos del BCRA y los parsea con openpyxl: stock_otros (series.xlsm,
hoja INSTRUMENTOS DEL BCRA col D) e itcrm_topup (ITCRMSerie.xlsx, hoja ITCRM y bilaterales).
NO incluye fx_al30 (viene de 1816) ni tcr_hist (base congelada).

Uso:  python sync_bcra.py            # produce macro/*.json
      python sync_bcra.py --check    # imprime cierres para validar
Metodologia identica a la aplicada a mano (ver RUNBOOK_DATOS.md del proyecto API 1816).
"""
from __future__ import annotations
import os, io, json, time, calendar, argparse
from datetime import date, datetime, timedelta
import urllib.request

BASE = "https://api.bcra.gob.ar/estadisticas/v4.0/monetarias/"
XLSX_BASE = "https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "macro")
os.makedirs(OUT, exist_ok=True)
HDR = {"User-Agent": "Mozilla/5.0 (bcra-sync)"}

# ---------------------------------------------------------------- fetch
_CACHE: dict = {}
def serie(idv, desde="2025-06-01", hasta=None):
    """Devuelve {fecha_str: valor} para un id de la API v4 (cachea por (id, desde)).
    Pagina con offset de a 3000 (la API topea en 3000 por pagina; CER id 30 es diario
    por dia calendario y supera 3000 filas desde 2015 -> requiere varias paginas)."""
    key = (idv, desde)
    if key in _CACHE:
        return _CACHE[key]
    hasta = hasta or date.today().isoformat()
    PAGE = 3000
    out: dict = {}
    offset = 0
    while True:
        url = f"{BASE}{idv}?desde={desde}&hasta={hasta}&limit={PAGE}&offset={offset}"
        got = 0
        for intento in range(4):
            try:
                req = urllib.request.Request(url, headers=HDR)
                with urllib.request.urlopen(req, timeout=60) as r:
                    j = json.loads(r.read().decode("utf-8"))
                det = (j.get("results") or [{}])[0].get("detalle", [])
                for d in det:
                    out[d["fecha"]] = d["valor"]
                got = len(det)
                break
            except Exception:
                if intento == 3:
                    raise
                time.sleep(2 * (intento + 1))
        if got < PAGE:
            break
        offset += PAGE
    _CACHE[key] = out
    return out

def last_common(ids):
    sets = [set(serie(i).keys()) for i in ids]
    return max(set.intersection(*sets))

def ref(d, target):
    t = target.isoformat() if isinstance(target, date) else target
    best = None
    for f in sorted(d):
        if f <= t: best = f
        else: break
    return d.get(best) if best else None

def ref_date(d, target):
    t = target.isoformat() if isinstance(target, date) else target
    best = None
    for f in sorted(d):
        if f <= t: best = f
        else: break
    return best

# ---------------------------------------------------------------- fechas EDATE
def _edate(d, months):
    m = d.month - 1 + months
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))

def edate_targets(cierre):
    mtd = date(cierre.year, cierre.month, 1) - timedelta(days=1)
    return [cierre, cierre - timedelta(days=7), mtd, _edate(cierre, -1),
            date(cierre.year - 1, 12, 31), _edate(cierre, -12)]

def niveles(ids, cierre):
    T = edate_targets(cierre)
    return {str(i): [ref(serie(i), t) for t in T] for i in ids}

def fechas_ref(id_probe, cierre):
    return [ref_date(serie(id_probe), t) for t in edate_targets(cierre)]

def flujo(idv, lo, hi):
    lo_s, hi_s = lo.isoformat(), hi.isoformat()
    return sum(v for f, v in serie(idv).items() if lo_s < f <= hi_s and v is not None)

# ---------------------------------------------------------------- nombres
N_TASAS = {44: "TAMAR bancos privados", 7: "BADLAR bancos privados", 8: "TM20 bancos privados",
           13: "Adelantos en cta. cte.", 14: "Prestamos personales", 1196: "Dep. a la vista remunerados"}
N_FACT = {47: "Compras divisas (s. priv. y otros)", 48: "Compras divisas (Tesoro)",
          50: "Transf. utilidades (Tesoro)", 51: "Resto operaciones (Tesoro)", 58: "LEFI",
          59: "Otros", 46: "TOTAL variacion BM"}
N_USD = {107: "Depositos USD (pub. + priv.)", 108: "Depositos USD privados",
         125: "Prestamos USD al s. privado", 118: "Adelantos", 119: "Documentos",
         120: "Hipotecarios", 121: "Prendarios", 122: "Personales", 123: "Tarjetas", 124: "Otros"}
N_RESF = {78: "Compra de divisas", 79: "Organismos internacionales", 80: "Otras oper. sector publico",
          81: "Efectivo minimo", 82: "Otras operaciones"}

def w(name, obj):
    json.dump(obj, open(os.path.join(OUT, name), "w", encoding="utf-8"), ensure_ascii=False)
    print("  ->", name)

# ---------------------------------------------------------------- bloques
def build_bcra_series():
    cierre = last_common([5])
    def ser(idv):
        d = serie(idv, desde="2025-12-30")
        return [[f, d[f]] for f in sorted(d) if f >= "2025-12-30"]
    w("bcra_series.json", {"meta": {"fuente": "api.bcra.gob.ar/estadisticas/v4.0/monetarias",
                                    "ids": {"oficial": 5, "techo": 1188, "piso": 1187},
                                    "nota": "oficial/techo/piso: diario YtD + ult. dato"},
                           "oficial": ser(5), "techo": ser(1188), "piso": ser(1187)})
    return cierre

def build_monetario():
    ids = [15, 17, 85, 91, 94, 95, 100, 109, 197]
    cierre = datetime.fromisoformat(last_common(ids)).date()
    w("monetario.json", {"meta": {"fuente": "api.bcra.gob.ar/estadisticas/v4.0/monetarias (saldos diarios, millones ARS)",
                                  "cierre_datos": cierre.isoformat(), "refs": ["ult", "WoW", "MtD", "MoM", "YTD", "YoY"],
                                  "fechas_ref": fechas_ref(15, cierre),
                                  "ids": {"15": "Base monetaria", "17": "Billetes y monedas en poder del publico",
                                          "85": "CC pub+priv", "91": "Dep pesos pub+priv (incl cedros)", "94": "CC privado",
                                          "95": "CA privado", "100": "Dep privado (incl cedros)", "109": "M2",
                                          "197": "M2 transaccional privado"}},
                         "series": niveles(ids, cierre)})

def build_monetario_ext():
    dep_ids = [85, 86, 87, 88, 89, 91, 94, 95, 96, 97, 100]
    ta_ids = [44, 7, 8, 13, 14, 1196]
    fac_ids = [47, 48, 50, 51, 58, 59, 46]
    pr_ids = [110, 111, 112, 113, 114, 115, 116, 117]
    depC = datetime.fromisoformat(last_common(dep_ids)).date()
    taC = datetime.fromisoformat(last_common(ta_ids)).date()
    facC = datetime.fromisoformat(max(serie(46))).date()
    prC = datetime.fromisoformat(last_common(pr_ids)).date()
    Tf = edate_targets(facC)
    win = {"1W": Tf[1], "1M": Tf[3], "YTD": date(facC.year - 1, 12, 31), "1Y": Tf[5]}
    fac = {str(i): {"nombre": N_FACT[i], "vals": [round(flujo(i, win[wn], facC)) for wn in ("1W", "1M", "YTD", "1Y")]}
           for i in fac_ids}
    s150, s151 = serie(150), serie(151)
    paC = datetime.fromisoformat(max(s150)).date()
    Tp = edate_targets(paC)
    tasa_refs = [ref(s150, t) for t in Tp]
    monto_refs = [None if k == 5 else round(ref(s151, Tp[k]) / 1e6, 3) for k in range(6)]
    ytd0 = "2026-01-01"
    tasa_serie = [[f, s150[f]] for f in sorted(s150) if f >= ytd0]
    monto_serie = [[f, round(s151[f] / 1e6, 3)] for f in sorted(s151) if f >= ytd0]
    w("monetario_ext.json", {"meta": {"fuente": "API BCRA v4 (saldos y tasas diarias; flujos de factores de la BM)",
                                      "dep": {"cierre": depC.isoformat(), "refs": ["ult", "WoW", "MtD", "MoM", "YTD", "YoY"],
                                              "fechas_ref": fechas_ref(85, depC)},
                                      "tasas": {"cierre": taC.isoformat(), "fechas_ref": fechas_ref(44, taC)},
                                      "factores": {"cierre": facC.isoformat(), "ventanas": ["1W", "1M", "YTD", "1Y"]}},
                             "depositos": niveles(dep_ids, depC),
                             "tasas": {str(i): {"nombre": N_TASAS[i], "vals": niveles([i], taC)[str(i)]} for i in ta_ids},
                             "factores": fac,
                             "prestamos": niveles(pr_ids, prC),
                             "pases_terceros": {"meta": {"tasa_id": 150, "monto_id": 151, "cierre": paC.isoformat(),
                                                         "fechas_ref": [ref_date(s150, t) for t in Tp]},
                                                "tasa_serie": tasa_serie, "monto_serie": monto_serie,
                                                "tasa_refs": tasa_refs, "monto_refs": monto_refs}})

def build_reservas():
    s78, s5 = serie(78), serie(5)
    fechas = sorted(f for f in s78 if f >= "2025-12-30")
    cum = 0.0; pes = 0.0; ser = []
    for f in fechas:
        v = s78.get(f) or 0
        cum += v; pes += v * (s5.get(f) or 0)
        ser.append([f, round(cum, 1), round(pes, 1)])
    w("reservas.json", {"meta": {"fuente": "API BCRA v4 id 78 (compra divisas) + A3500 id 5 (pesos)",
                                 "cierre_datos": fechas[-1], "nota": "cumsum id78; pesos = compra x A3500"},
                        "series": ser})
    fac_ids = [78, 79, 80, 81, 82]
    facC = datetime.fromisoformat(max(serie(78))).date()
    Tf = edate_targets(facC)
    win = {"1W": Tf[1], "1M": Tf[3], "YTD": date(facC.year - 1, 12, 31), "1Y": Tf[5]}
    fac = {str(i): {"nombre": N_RESF[i], "vals": [round(flujo(i, win[wn], facC)) for wn in ("1W", "1M", "YTD", "1Y")]}
           for i in fac_ids}
    fac["TOT"] = {"nombre": "TOTAL variacion reservas", "vals": [sum(fac[str(i)]["vals"][k] for i in fac_ids) for k in range(4)]}
    s1 = serie(1)
    stock = [[f, s1[f]] for f in sorted(s1) if f >= "2025-12-30"]
    w("reservas_ext.json", {"meta": {"fuente": "API BCRA v4: stock id 1; factores 78-82",
                                     "ventanas": ["1W", "1M", "YTD", "1Y"], "stock_ult": stock[-1]},
                            "stock_serie": stock, "factores": fac})

def _prom_fin(s):
    from collections import defaultdict
    by = defaultdict(list)
    for f, v in s.items():
        if v is not None and f >= "2015-01-01": by[f[:4]].append((f, v))
    out = {}
    for y, it in by.items():
        it.sort(); vals = [v for _, v in it]
        out[y] = (sum(vals) / len(vals), it[-1][1])
    return out

def _anual_usd():
    s125 = serie(125, desde="2015-01-01"); s108 = serie(108, desde="2015-01-01")
    cred = _prom_fin(s125); dep = _prom_fin(s108)
    years = sorted(set(cred) & set(dep))
    return {"nota": "promedio anual diario y fin de anio, ids 125 (creditos) y 108 (depositos priv)",
            "cred": [[y, round(cred[y][0]), round(cred[y][1])] for y in years],
            "dep": [[y, round(dep[y][0]), round(dep[y][1])] for y in years]}

def _anual_prestamos():
    s117 = serie(117, desde="2015-01-01"); s100 = serie(100, desde="2015-01-01"); scer = serie(30, desde="2015-01-01")
    base_d = max(scer); cer_base = scer[base_d]
    from collections import defaultdict
    by117 = defaultdict(list); by100 = defaultdict(list)
    for f, v in s117.items():
        if v is not None and f >= "2015-01-01": by117[f[:4]].append((f, v))
    for f, v in s100.items():
        if v is not None and f >= "2015-01-01": by100[f[:4]].append((f, v))
    years = sorted(set(by117) & set(by100))
    filas = []
    for y in years:
        it = sorted(by117[y])
        reales = [v * cer_base / (ref(scer, f) or cer_base) for f, v in it]
        prom_real = sum(reales) / len(reales) / 1e6      # millones ARS -> billones (Bn)
        fin_real = reales[-1] / 1e6
        vals100 = [v for _, v in sorted(by100[y])]
        ratio = (sum(v for _, v in it) / len(it)) / (sum(vals100) / len(vals100))
        filas.append([y, round(prom_real, 2), round(fin_real, 2), round(ratio, 4)])
    return {"nota": "Prestamos $ al s. privado (id 117) deflactados por CER (id 30) a pesos del cierre.",
            "base_date": base_d, "cer_base": round(cer_base, 3), "filas": filas}

def build_usd_cer():
    usd_ids = [107, 108, 125, 118, 119, 120, 121, 122, 123, 124]
    depC = datetime.fromisoformat(last_common([108, 125])).date()
    usd = {str(i): {"nombre": N_USD[i], "vals": niveles([i], depC)[str(i)]} for i in usd_ids}
    s125, s108, s30 = serie(125), serie(108), serie(30, desde="2025-06-01")
    ytd0 = "2025-12-30"
    cred_serie = [[f, s125[f]] for f in sorted(s125) if f >= ytd0]
    dep_serie = [[f, s108[f]] for f in sorted(s108) if f >= ytd0]
    cerC = datetime.fromisoformat(max(s30)).date()
    Tc = edate_targets(cerC); Td = edate_targets(depC)
    cer = {"ult": round(ref(s30, cerC), 4),
           "refs_propios": [round(ref(s30, t), 4) for t in Tc],
           "en_grilla_dep": [round(ref(s30, ref_date(serie(108), t)), 4) for t in Td],
           "corridos_45d": {"2025-08-14": round(ref(s30, date(2025, 8, 14)), 4),
                            "2026-02-13": round(ref(s30, date(2026, 2, 13)), 4),
                            "nota": "CER en t0+45d (YTD t0=30-dic, YoY t0=30-jun-25); capado y prorrateado"}}
    w("usd_cer.json", {"meta": {"fuente": "API BCRA v4. USD ids 107/108/125/118-124; CER id 30.",
                                "dep_refs": fechas_ref(108, depC), "cer_refs": [ref_date(s30, t) for t in Tc]},
                       "usd": usd, "cred_serie": cred_serie, "dep_serie": dep_serie, "cer": cer,
                       "anual_usd": _anual_usd(), "anual_prestamos": _anual_prestamos()})

def build_bm_componentes():
    s15, s16 = serie(15, desde="2015-01-01"), serie(16, desde="2015-01-01")
    from collections import defaultdict
    ratios = defaultdict(list)
    for f in sorted(set(s15) & set(s16)):
        if s15[f]: ratios[f[:7]].append(s16[f] / s15[f])
    share = [[ym, round(sum(v) / len(v), 3)] for ym, v in sorted(ratios.items())]
    w("bm_componentes.json", {"meta": {"fuente": "API BCRA v4 ids 15 (BM) y 16 (circulacion)",
                                       "nota": "share circulacion mensual (prom. de ratios diarios)",
                                       "ult_dato": share[-1][0]},
                              "share_circ_mensual": share})

# ---------------------------------------------------------------- xlsx del BCRA (openpyxl)
NOTA13 = ("Comprende el saldo del pasivo (con signo positivo) del BCRA neto del activo "
          "(con signo negativo) para el BCRA, resultante de operaciones de corto plazo, "
          "concertadas a partir del 15 de julio de 2025, tomando como referencia las tasas "
          "de interes de mercado vigentes. Incluye Operaciones Simultaneas negociadas en BYMA.")

def _download(fname):
    url = XLSX_BASE + fname
    for intento in range(4):
        try:
            req = urllib.request.Request(url, headers=HDR)
            with urllib.request.urlopen(req, timeout=120) as r:
                return r.read()
        except Exception:
            if intento == 3:
                raise
            time.sleep(3 * (intento + 1))

def _as_date(v):
    """Normaliza una celda de fecha a 'YYYY-MM-DD' (datetime o serial Excel)."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool) and 30000 < v < 80000:
        return (date(1899, 12, 30) + timedelta(days=int(round(v)))).isoformat()
    return None

def build_stock_otros():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(_download("series.xlsm")), data_only=True, keep_vba=False)
    ws = wb["INSTRUMENTOS DEL BCRA"]
    serie = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        f = _as_date(row[0]); v = row[3]
        if f and f >= "2025-12-30" and isinstance(v, (int, float)) and not isinstance(v, bool):
            serie.append([f, round(v)])
    wb.close()
    serie.sort()
    w("stock_otros.json", {"meta": {"fuente": "series.xlsm BCRA, hoja 'INSTRUMENTOS DEL BCRA', col D 'Otros (13)' (saldos, millones ARS).",
                                    "nota13": NOTA13, "actualizado": date.today().isoformat(),
                                    "ult_dato": serie[-1][0] if serie else None},
                          "serie": serie})
    return serie[-1] if serie else None

def build_itcrm():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(_download("ITCRMSerie.xlsx")), data_only=True)
    ws = wb["ITCRM y bilaterales"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        f = _as_date(row[0]); v = row[1]
        if f and isinstance(v, (int, float)) and not isinstance(v, bool):
            rows.append([f, round(v, 6)])
    wb.close()
    rows.sort()
    w("itcrm_topup.json", {"meta": {"fuente": "ITCRMSerie.xlsx (BCRA), hoja 'ITCRM y bilaterales'",
                                    "actualizado": date.today().isoformat(),
                                    "nota": "incluye dias no habiles; el merge filtra por fechas con TCN; ultimas 60 filas para captar revisiones"},
                          "itcrm": rows[-60:]})
    return rows[-1] if rows else None

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--check", action="store_true"); a = ap.parse_args()
    print("Bajando y reduciendo BCRA API v4 ->", OUT)
    print("bcra_series cierre", build_bcra_series())
    build_monetario(); build_monetario_ext(); build_reservas(); build_usd_cer(); build_bm_componentes()
    print("stock_otros ult", build_stock_otros())
    print("itcrm ult", build_itcrm())
    print("Listo.")
    if a.check:
        for n in ("bcra_series", "monetario", "monetario_ext", "reservas", "reservas_ext", "usd_cer",
                  "bm_componentes", "stock_otros", "itcrm_topup"):
            d = json.load(open(os.path.join(OUT, n + ".json"))); m = d.get("meta", {})
            dep = m.get("dep") or {}
            c = (m.get("cierre_datos") or dep.get("cierre") or m.get("ult_dato")
                 or (m.get("stock_ult") or [""])[0] or m.get("actualizado") or "")
            print("  %s cierre=%s" % (n, c))

if __name__ == "__main__":
    main()
