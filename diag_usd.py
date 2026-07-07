#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Diagnostico one-shot: baja la planilla mas reciente y muestra como vienen los fondos USD.
No commitea nada. Se corre por el workflow diag.yml."""
import io
import collections
import openpyxl
import sync_cnv as S


def main():
    docs = S.listar_documentos()
    fecha, pres = docs[0]
    print("DOC", fecha, pres)
    fg = S.file_guid(pres)
    content = S.descargar_xlsx(fg[1], fg[2])
    if not content:
        print("NO CONTENT"); return
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Sheet 1"] if "Sheet 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    print("SHEETS", wb.sheetnames)

    # cabeceras (primeras filas) para confirmar significado de columnas
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        cells = [(j, str(v)[:16]) for j, v in enumerate(r) if v is not None]
        print("HDR", i + 1, cells[:22])

    allrows = list(ws.iter_rows(min_row=9, values_only=True))
    mon = collections.Counter()
    clasifs = collections.Counter()
    usd_examples = []
    clasif = None
    for r in allrows:
        c0 = r[0]
        vcp = r[5] if len(r) > 5 else None
        patr = r[14] if len(r) > 14 else None
        if c0 and vcp is None and patr is None:
            clasif = str(c0).strip(); clasifs[clasif] += 0; continue
        if not c0:
            continue
        if clasif:
            clasifs[clasif] += 1
        m = r[1] if len(r) > 1 else None
        mon[str(m)] += 1
        txt = (str(m) + " " + str(clasif)).upper()
        if ("OLAR" in txt or "USD" in txt or "U$" in txt or "DOL" in txt):
            if len(usd_examples) < 5:
                usd_examples.append((clasif, list(enumerate(r))))

    print("MONEDA_COUNTS", dict(mon))
    print("CLASIF_HEADERS", dict(clasifs))
    for cl, row in usd_examples:
        print("==== USDROW clasif=", cl)
        for j, v in row:
            if v is not None:
                print("   col", j, "=", str(v)[:44])
    wb.close()


if __name__ == "__main__":
    main()
