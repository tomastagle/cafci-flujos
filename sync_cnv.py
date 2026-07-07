#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sincronizador de FLUJOS de FCI desde CNV (oficial, publico).  v5.1 (07-jul-2026)

Cambios v4 vs v3:
  1) MONEDA: los fondos en dolares (moneda USD/USB o clasif "Dolar Estadounidense") se separan
     en buckets con sufijo " USD" en TODAS las categorias (antes solo MM y Renta Fija). Sus valores
     ya vienen en USD en la planilla (col5 VCP y col14 Patrimonio en la moneda del fondo), asi que
     no se mezclan mas con los pesos. El reporte los muestra en una tabla aparte en USD.
  2) GUARD DE ARTEFACTO: el flujo diario solo se computa si |flujo| <= 0.85 * patrimonio del fondo
     (max entre hoy y dia previo). Los saltos ~100% (altas/bajas/reclasificacion de clases, cuando
     falta la cuotaparte del dia previo) se descartan: eran los que contaminaban YTD/1Y.
  3) REVISIONES: en modo diario (backfill 0) se re-bajan los ultimos 8 dias habiles para captar
     re-presentaciones de CNV (la mas nueva por fecha = ultima revision).

Se mantiene de v3: historico agregado por dia (cnv_agg_history.csv) y el mismo JSON de salida.

Circuito de descarga (sin token, publico):
  1) GET  www.cnv.gov.ar/.../CuotaPartes  -> (fecha_doc, GUID presentacion del link al AIF)
  2) GET  aif2.cnv.gov.ar/Presentations/publicview/{GUID} -> HTML con {nombreArchivo, guid} + token CfDJ8
  3) GET  aif2.cnv.gov.ar/api/ValetKeyProvider/GetPublicValetKey/{fileGuid}?operation=DownloadBlob
  4) POST blob.cnv.gov.ar/BlobWebService.svc/DownloadBlob/{fileGuid}  (X-CSRF-TOKEN + body JSON {"ValetKey":...})

Uso: python sync_cnv.py [--backfill N] [--date AAAA-MM-DD]
"""
import os
import re
import io
import sys
import csv
import json
import time
import argparse
import datetime as dt

try:
    from curl_cffi import requests as _rq
    _IMPERSONATE = "chrome"
except Exception:
    import requests as _rq
    _IMPERSONATE = None

import openpyxl

CNV_LIST = "https://www.cnv.gov.ar/SitioWeb/FondosComunesInversion/CuotaPartes"
AIF_VIEW = "https://aif2.cnv.gov.ar/Presentations/publicview/"
AIF_VALET = "https://aif2.cnv.gov.ar/api/ValetKeyProvider/GetPublicValetKey/"
BLOB_DL = "https://blob.cnv.gov.ar/BlobWebService.svc/DownloadBlob/"

HEADERS = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
           "Accept-Language": "es-AR,es;q=0.9"}
SESS = _rq.Session()

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
AGG = os.path.join(OUT_DIR, "cnv_agg_history.csv")
AGG_COLS = ["fecha", "dim", "clave", "in", "out", "net", "aum"]
DIMS = ("bucket", "gestora", "tipo_ger")
BM_HIST = os.path.join(OUT_DIR, "bm_hist.csv")
BM_COLS = ["fecha", "key", "nombre", "tipo", "moneda", "vcp", "vcp_prev", "aum"]
BM_KEY = "bull market"   # gestora de los 6 FCIs de Bull Market

# umbral del guard de artefacto (fraccion del patrimonio del fondo)
MAX_FLUJO_FRAC = 0.85

_MESES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
          "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}

# gestoras BANCARIAS (keywords, ex managerType de fonditos: BANK). El resto = Independiente.
BANCARIAS = ["santander", "galicia", "bbva", "icbc", "macro", "mariva", "patagonia",
             "supervielle", "bancor", "provinfondos", "pellegrini", "industrial",
             "bacs", "ciudad", "nacion", "nación"]


# ------------------------------- HTTP ------------------------------------------
def _req(method, url, tries=5, backoff=2.5, **kw):
    kw.setdefault("timeout", 90); kw.setdefault("headers", HEADERS)
    if _IMPERSONATE:
        kw["impersonate"] = _IMPERSONATE
    for i in range(tries):
        try:
            r = SESS.request(method, url, **kw)
            if r.status_code == 200:
                return r
            print("   ! HTTP %s %s" % (r.status_code, url), file=sys.stderr)
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(backoff * (i + 1)); continue
            return None
        except Exception as e:
            print("   ! %s (%s)" % (str(e)[:120], url), file=sys.stderr)
            time.sleep(backoff * (i + 1))
    return None


def listar_documentos():
    r = _req("GET", CNV_LIST)
    if not r:
        raise SystemExit("No se pudo leer la lista de CuotaPartes de CNV.")
    pat = re.compile(r'publicview/([0-9A-Fa-f-]{30,})"\s*>\s*(\d{1,2})\s+([a-z]{3})\.?\s+(\d{4})', re.I)
    out = []
    for m in pat.finditer(r.text):
        d, mon, y = int(m.group(2)), _MESES.get(m.group(3).lower()), int(m.group(4))
        if mon:
            out.append(("%04d-%02d-%02d" % (y, mon, d), m.group(1)))
    return out


def file_guid(pres_guid):
    r = _req("GET", AIF_VIEW + pres_guid)
    if not r:
        return None
    html = r.text
    m = re.search(r'"nombreArchivo":"([^"]*Planilla_Diaria[^"]*\.xlsx)"\s*,\s*"tamano":"[^"]*"\s*,\s*"guid":"([0-9a-f-]+)"',
                  html, re.I) or re.search(r'"nombreArchivo":"([^"]*\.xlsx)"[^{}]*?"guid":"([0-9a-f-]+)"', html, re.I)
    if not m:
        return None
    tm = re.search(r'RequestVerificationToken"[^>]*\bvalue="([^"]+)"', html) or re.search(r'value="(CfDJ8[^"]+)"', html)
    return (m.group(1), m.group(2), tm.group(1) if tm else "")


def _valet_key(fguid):
    r = _req("GET", AIF_VALET + fguid + "?operation=DownloadBlob")
    try:
        return r.json().get("valetKeyData") if r else None
    except Exception:
        return None


def descargar_xlsx(fguid, csrf=""):
    vk = _valet_key(fguid)
    if not vk:
        print("   ! sin valet key", file=sys.stderr); return None
    url = BLOB_DL + fguid
    hdr = dict(HEADERS, **{"X-CSRF-TOKEN": csrf, "Content-Type": "application/json",
                           "Referer": "https://aif2.cnv.gov.ar/"})
    r = _req("POST", url, headers=hdr, data=json.dumps({"ValetKey": vk}))
    if not (r and r.content[:2] == b"PK"):
        r = _req("POST", url, headers=dict(HEADERS, **{"X-CSRF-TOKEN": csrf, "Referer": "https://aif2.cnv.gov.ar/"}),
                 data={"ValetKey": vk})
    return r.content if (r and r.content[:2] == b"PK") else None


# ------------------------------- parseo + flujo intra-archivo ------------------
def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def tipo_gerente(g):
    gl = (g or "").lower()
    return "Bancaria" if any(k in gl for k in BANCARIAS) else "Independiente"


def es_usd(clasif, moneda):
    c = (clasif or "").upper()
    return ("DOLAR" in c) or ((moneda or "").upper().strip() in ("USD", "USB", "U$S", "DOL"))


def bucket(clasif, plazo, moneda):
    c = (clasif or "").upper()
    usd = es_usd(clasif, moneda)
    try:
        p = int(plazo)
    except (TypeError, ValueError):
        p = None
    if "MERCADO DE DINERO" in c:
        return "Money Market USD" if usd else "Money Market ARS (T+0)"
    if "RENTA FIJA" in c:
        if usd:
            return "Renta Fija USD"
        return "Renta Fija ARS T+1" if p == 1 else "Renta Fija ARS"
    if "RENTA MIXTA" in c:
        return "Renta Mixta USD" if usd else "Renta Mixta"
    if "RENTA VARIABLE" in c:
        return "Renta Variable USD" if usd else "Renta Variable"
    if "PYME" in c:
        return "PyMEs USD" if usd else "PyMEs"
    if "INFRAEST" in c:
        return "Infraestructura USD" if usd else "Infraestructura"
    if "RETORNO TOTAL" in c:
        return "Retorno Total USD" if usd else "Retorno Total"
    if "ASG" in c:
        return "ASG USD" if usd else "ASG"
    if "RG900" in c:
        return "RG900"
    if "FONDOS CERRADOS" in c:
        return "Fondos Cerrados USD" if usd else "Fondos Cerrados"
    base = (clasif or "Otros").replace(" Peso Argentina", "") \
        .replace(" Dolar Estadounidense Billete", " USD").replace(" Dolar Estadounidense", " USD")
    return base


def parse_planilla(content):
    """Devuelve (fecha_iso, [filas por fondo con FLUJO DIARIO intra-archivo, con guard de artefacto])."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Sheet 1"] if "Sheet 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=9, values_only=True))
    wb.close()
    clasif = None
    fecha_iso = None
    out = []
    bm = []
    for r in rows:
        c0 = r[0]
        vcp = _num(r[5]); patr = _num(r[14]); cn = _num(r[12]); cp = _num(r[13])
        patr_prev = _num(r[15]) if len(r) > 15 else None
        if c0 and vcp is None and patr is None:
            clasif = str(c0).strip(); continue
        if not c0 or vcp is None:
            continue
        if fecha_iso is None and r[4]:
            for fmt in ("%d/%m/%y", "%d/%m/%Y"):
                try:
                    fecha_iso = dt.datetime.strptime(str(r[4]).strip(), fmt).strftime("%Y-%m-%d"); break
                except Exception:
                    pass
        # flujo intra-archivo con guard de artefacto (descartar saltos ~totales de patrimonio)
        flujo = None
        if cn is not None and cp is not None:
            f = (cn - cp) * vcp / 1000.0
            ref = max(abs(patr or 0.0), abs(patr_prev or 0.0))
            if ref > 0 and abs(f) <= MAX_FLUJO_FRAC * ref:
                flujo = f
        ger = str(r[23]).strip() if r[23] else "?"
        bk = bucket(clasif, r[37], r[1])
        out.append(dict(fondo=str(c0).strip(), gestora=ger, tipo_ger=tipo_gerente(ger),
                        bucket=bk, aum=patr, flujo=flujo))
        if BM_KEY in ger.lower():
            bm.append(dict(nombre=str(c0).strip(), tipo=bk,
                           moneda=(str(r[1]).strip() if r[1] else ""),
                           vcp=vcp, vcp_prev=_num(r[6]), aum=patr))
    return fecha_iso, out, bm


def agregar_dia(filas):
    """De filas por fondo -> filas agregadas (dim, clave, in, out, net, aum)."""
    acc = {}
    for f in filas:
        fl = f["flujo"]; a = f["aum"] or 0.0
        for dim in DIMS:
            clave = f.get(dim) or "?"
            d = acc.setdefault((dim, clave), [0.0, 0.0, 0.0, 0.0])
            if fl is not None:
                if fl >= 0:
                    d[0] += fl
                else:
                    d[1] += fl
                d[2] += fl
            d[3] += a
    return [(dim, clave, v[0], v[1], v[2], v[3]) for (dim, clave), v in acc.items()]


# ------------------------------- history (agregado) ----------------------------
def append_agg(fecha_iso, aggrows):
    os.makedirs(OUT_DIR, exist_ok=True)
    prev = []
    if os.path.exists(AGG):
        with open(AGG, newline="", encoding="utf-8") as fh:
            prev = [r for r in csv.DictReader(fh) if r["fecha"] != fecha_iso]
    with open(AGG, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=AGG_COLS); w.writeheader()
        for r in prev:
            w.writerow({k: r.get(k, "") for k in AGG_COLS})
        for (dim, clave, i, o, n, a) in aggrows:
            w.writerow({"fecha": fecha_iso, "dim": dim, "clave": clave,
                        "in": "%.2f" % i, "out": "%.2f" % o, "net": "%.2f" % n, "aum": "%.2f" % a})
    print("   agg: %d filas para %s (revisiones reemplazadas)" % (len(aggrows), fecha_iso))


def load_agg():
    if not os.path.exists(AGG):
        return []
    with open(AGG, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for r in rows:
        for k in ("in", "out", "net", "aum"):
            r[k] = _num(r[k]) or 0.0
    return rows


# ------------------------------- flujos (ventanas = suma de diarios) -----------
def append_bm(fecha_iso, bmrows):
    if not bmrows:
        return
    os.makedirs(OUT_DIR, exist_ok=True)
    prev = []
    if os.path.exists(BM_HIST):
        with open(BM_HIST, newline="", encoding="utf-8") as fh:
            prev = [r for r in csv.DictReader(fh) if r["fecha"] != fecha_iso]
    with open(BM_HIST, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=BM_COLS); w.writeheader()
        for r in prev:
            w.writerow({k: r.get(k, "") for k in BM_COLS})
        for b in bmrows:
            w.writerow({"fecha": fecha_iso, "key": b["nombre"], "nombre": b["nombre"], "tipo": b["tipo"],
                        "moneda": b["moneda"], "vcp": b["vcp"], "vcp_prev": b["vcp_prev"], "aum": b["aum"]})


def load_bm():
    if not os.path.exists(BM_HIST):
        return []
    with open(BM_HIST, newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for r in rows:
        for k in ("vcp", "vcp_prev", "aum"):
            r[k] = _num(r[k])
    return rows


def _base_fondo(n):
    return re.split(r"\s*[-\u2013]?\s*Clase\b", n, flags=re.I)[0].strip()


def build_bm_funds(cierre):
    rows = load_bm()
    if not rows:
        return []
    fechas = sorted({r["fecha"] for r in rows})
    if cierre not in fechas:
        cierre = fechas[-1]
    vcp_by = {(r["key"], r["fecha"]): r["vcp"] for r in rows}
    fset = sorted(dt.date.fromisoformat(f) for f in fechas)
    c = dt.date.fromisoformat(cierre)
    idx = fset.index(c) if c in fset else len(fset) - 1

    def ref_date(target):
        prev = [f for f in fset if f <= target]
        return prev[-1].isoformat() if prev else None

    starts = {"1D": fset[idx - 1].isoformat() if idx > 0 else None,
              "WTD": ref_date(c - dt.timedelta(days=7)),
              "MTD": ref_date(dt.date(c.year, c.month, 1) - dt.timedelta(days=1)),
              "YTD": ref_date(dt.date(c.year, 1, 1) - dt.timedelta(days=1)),
              "1Y": ref_date(c - dt.timedelta(days=365))}
    latest = [r for r in rows if r["fecha"] == cierre]
    grupos = {}
    for r in latest:
        grupos.setdefault(_base_fondo(r["nombre"]), []).append(r)
    out = []
    for base, clases in grupos.items():
        rep = max(clases, key=lambda r: (r["aum"] or 0))
        aum_tot = sum((r["aum"] or 0) for r in clases)
        vcp = rep["vcp"]; key = rep["key"]

        def ret(win):
            if win == "1D" and rep.get("vcp_prev"):
                ref = rep["vcp_prev"]
            else:
                f0 = starts.get(win)
                ref = vcp_by.get((key, f0)) if f0 else None
            return (vcp / ref - 1) if (vcp and ref) else None

        r1d = ret("1D")
        out.append({"nombre": base, "tipo": rep["tipo"], "moneda": rep["moneda"],
                    "vcp": vcp, "aum": aum_tot, "es_mm": rep["tipo"].startswith("Money Market"),
                    "tna": (r1d * 365 if r1d is not None else None),
                    "ret": {"1D": r1d, "WTD": ret("WTD"), "MTD": ret("MTD"),
                            "YTD": ret("YTD"), "1Y": ret("1Y")}})
    out.sort(key=lambda z: -(z["aum"] or 0))
    return out


WINDOWS = ["1D", "1W", "1M", "YTD", "1Y"]


def _starts(fechas, cierre):
    c = dt.date.fromisoformat(cierre)
    fset = sorted(dt.date.fromisoformat(f) for f in fechas)
    idx = fset.index(c) if c in fset else len(fset) - 1
    tgt = {"1W": c - dt.timedelta(days=7), "1M": c - dt.timedelta(days=30),
           "YTD": dt.date(c.year, 1, 1), "1Y": c - dt.timedelta(days=365)}
    res = {"1D": fset[idx - 1].isoformat() if idx > 0 else None}
    for w, t in tgt.items():
        prev = [f for f in fset if f <= t]
        res[w] = (prev[-1] if prev else fset[0]).isoformat()
    return res


def flujos(cierre):
    rows = load_agg()
    if not rows:
        return None
    fechas = sorted({r["fecha"] for r in rows})
    if cierre not in fechas:
        cierre = fechas[-1]
    starts = _starts(fechas, cierre)

    def agg_win(win, dim):
        f0 = starts.get(win)
        if not f0:
            return {}
        m = {}
        for r in rows:
            if r["dim"] != dim:
                continue
            if f0 < r["fecha"] <= cierre:
                d = m.setdefault(r["clave"], {"in": 0.0, "out": 0.0, "net": 0.0})
                d["in"] += r["in"]; d["out"] += r["out"]; d["net"] += r["net"]
        return dict(sorted(m.items(), key=lambda kv: kv[1]["net"], reverse=True))

    # serie diaria de neto de la industria (suma de net del dim 'bucket' por fecha), ult. 40
    dia = {}
    for r in rows:
        if r["dim"] == "bucket" and not r["clave"].endswith("USD"):
            dia[r["fecha"]] = dia.get(r["fecha"], 0.0) + r["net"]
    serie = [[f, round(dia[f])] for f in fechas if f in dia][-40:]

    def aum_dim(dim):
        m = {}
        for r in rows:
            if r["fecha"] == cierre and r["dim"] == dim:
                m[r["clave"]] = m.get(r["clave"], 0.0) + r["aum"]
        return dict(sorted(m.items(), key=lambda kv: -kv[1]))

    res = {"cierre": cierre, "ventanas_inicio": starts,
           "por_bucket": {w: agg_win(w, "bucket") for w in WINDOWS},
           "por_gestora": {w: agg_win(w, "gestora") for w in WINDOWS},
           "por_tipo_ger": {w: agg_win(w, "tipo_ger") for w in WINDOWS},
           "serie_diaria": serie,
           "aum_por_bucket": aum_dim("bucket"),
           "aum_por_gestora": aum_dim("gestora"),
           "aum_por_tipo_ger": aum_dim("tipo_ger"),
           "bm_funds": build_bm_funds(cierre)}
    return res


# ------------------------------- orquestacion ----------------------------------
def _bajar(pres):
    fg = file_guid(pres)
    if not fg:
        return None, None, None
    content = descargar_xlsx(fg[1], fg[2])
    if not content:
        return None, None, None
    return parse_planilla(content)


def bajar_un_dia(fecha_target):
    docs = listar_documentos()
    cand = [g for (f, g) in docs if f == fecha_target]
    if not cand:
        raise SystemExit("No hay doc para %s." % fecha_target)
    print("Documento %s (%s)" % (fecha_target, cand[0]))
    fecha_iso, filas, bm = _bajar(cand[0])
    if filas is None:
        raise SystemExit("No se pudo descargar/parsear el xlsx.")
    print("  %d fondos (fecha %s) | BM %d" % (len(filas), fecha_iso, len(bm)))
    append_agg(fecha_iso or fecha_target, agregar_dia(filas))
    append_bm(fecha_iso or fecha_target, bm)
    return fecha_iso or fecha_target


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date")
    ap.add_argument("--backfill", type=int, default=0)
    args = ap.parse_args()
    os.makedirs(OUT_DIR, exist_ok=True)

    if args.date:
        cierre = bajar_un_dia(args.date)
    else:
        # backfill 0 => refrescar ultimos 8 dias habiles (captura re-presentaciones/revisiones)
        n = args.backfill if args.backfill else 8
        docs = listar_documentos(); vistas = set()
        for (fecha_doc, pres) in docs:
            if fecha_doc in vistas:
                continue
            if len(vistas) >= n:
                break
            vistas.add(fecha_doc)
            try:
                fecha_iso, filas, bm = _bajar(pres)
                if filas is None:
                    continue
                append_agg(fecha_iso or fecha_doc, agregar_dia(filas))
                append_bm(fecha_iso or fecha_doc, bm)
                time.sleep(1.0)
            except Exception as e:
                print("  ! backfill %s: %s" % (fecha_doc, str(e)[:80]), file=sys.stderr)
        cierre = sorted(vistas)[-1] if vistas else None

    if not cierre:
        raise SystemExit("Sin cierre.")
    res = flujos(cierre)
    with open(os.path.join(OUT_DIR, "flujos_latest.json"), "w", encoding="utf-8") as fh:
        json.dump({"generado": dt.datetime.utcnow().isoformat() + "Z", "flujos": res}, fh, ensure_ascii=False)
    if res:
        neto = sum(v["net"] for b, v in res["por_bucket"]["1D"].items() if not b.endswith("USD"))
        print("OK -> flujos_latest.json (cierre %s | neto 1D ARS %+.0f mill)" % (res["cierre"], neto / 1e6))


if __name__ == "__main__":
    main()
